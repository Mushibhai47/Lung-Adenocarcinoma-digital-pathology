"""
build_master_table.py
Creates Complete_Data_Table.xlsx — all 172 patients, all clinical variables
decoded into readable labels, plus digital pathology columns for the 141 mapped cases.
Vania can open this directly in Excel or import into SPSS for further analysis.
"""
import sys, warnings
sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore')

import pyreadstat
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
VERY_LIGHT = "DEEAF1"
GREEN_FILL = "E2EFDA"
ORANGE     = "FCE4D6"
YELLOW     = "FFF2CC"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"

def thin_border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

# ── Load SPSS ─────────────────────────────────────────────────────────────
df_spss, meta = pyreadstat.read_sav(
    r'C:\Users\musha\Desktop\QuPath\Lung_june2025.sav')
vl = meta.variable_value_labels

# ── Decode value labels → human-readable ─────────────────────────────────
df = df_spss.copy()

def decode_col(df, col):
    if col in vl:
        return df[col].map(vl[col]).fillna(df[col])
    return df[col]

# Build decoded dataframe
decoded = pd.DataFrame()
decoded['FICHA']                       = df['FICHA']
decoded['Patient_sex']                 = decode_col(df, 'Patient_sex')
decoded['Patient_age']                 = pd.to_numeric(df['Patient_age'], errors='coerce').round(0).astype('Int64')
decoded['Surgery_date']                = pd.to_datetime(df['Surgery_date']).dt.strftime('%Y-%m-%d')
decoded['Last_contact_or_death_date']  = pd.to_datetime(df['dateofdeathorlastcontact']).dt.strftime('%Y-%m-%d')

# Survival
df['Surgery_date_dt'] = pd.to_datetime(df['Surgery_date'])
df['end_date_dt']     = pd.to_datetime(df['dateofdeathorlastcontact'])
df['surv_months']     = (df['end_date_dt'] - df['Surgery_date_dt']).dt.days / 30.44
decoded['Survival_months']             = df['surv_months'].round(1)
decoded['Status']                      = decode_col(df, 'Status')   # alive / dead

# Tumour
decoded['Tumor_focality']              = decode_col(df, 'Tumor_focality')
decoded['Tumor_laterality']            = decode_col(df, 'Tumor_laterality')
decoded['Tumor_location']              = decode_col(df, 'Tumor_location')
decoded['Tumor_size_cm']               = df['Tumor_size']
decoded['Surgical_procedure']          = decode_col(df, 'Surgical_procedure')
decoded['Margins']                     = decode_col(df, 'Margins')
decoded['Pathological_diagnosis']      = decode_col(df, 'Pathological_diagnosis')
decoded['Tumor_predominant_pattern']   = decode_col(df, 'Tumor_predominant_pattern')

# Staging
decoded['pT']                          = decode_col(df, 'pT')
decoded['pN']                          = decode_col(df, 'pN')
decoded['Stage_group']                 = decode_col(df, 'Stage_group')
decoded['Number_lymph_nodes_submitted']= pd.to_numeric(df['Number_lymph_nodes_submitted'], errors='coerce').round(0).astype('Int64')
decoded['Number_metastatic_LN']        = pd.to_numeric(df['Number_metastatic_lymph_nodes'], errors='coerce').round(0).astype('Int64')

# Histologic features
decoded['STAS']                        = decode_col(df, 'STAS')
decoded['Lymphovascular_invasion']     = decode_col(df, 'Lymphovascular_Invasion')
decoded['Visceral_pleural_invasion']   = decode_col(df, 'Visceral_Pleural_Invasion')

# Molecular
decoded['Molecular_method']            = decode_col(df, 'Molecular_method')
decoded['Molecular_results']           = decode_col(df, 'Molecular_results')

# IHC
decoded['PDL1_IHC']                    = decode_col(df, 'PDL1_IHQ')
decoded['Vimentin_IHC']                = decode_col(df, 'Vimentine_IHQ')

# ── Merge digital pathology data ──────────────────────────────────────────
df_dig = pd.read_csv(r'C:\Users\musha\Desktop\QuPath\analysis_results.csv')

# Keep only the digital columns we want to add
dig_cols = ['FICHA', 'slide_id',
            'lymp_ROI1', 'lymp_ROI2', 'lymp_mean',
            'tumor_ROI1', 'tumor_ROI2', 'tumor_mean',
            'ratio_mean']
df_dig_slim = df_dig[dig_cols].copy()
df_dig_slim.columns = ['FICHA', 'Slide_ID',
                        'Lymp_density_ROI1 (cells/mm2)',
                        'Lymp_density_ROI2 (cells/mm2)',
                        'Lymp_density_MEAN (cells/mm2)',
                        'Tumor_density_ROI1 (cells/mm2)',
                        'Tumor_density_ROI2 (cells/mm2)',
                        'Tumor_density_MEAN (cells/mm2)',
                        'Tumor_Lymp_ratio']

# Round to 1 decimal
for c in df_dig_slim.columns[2:]:
    df_dig_slim[c] = df_dig_slim[c].round(1)

# Merge (left join keeps all 172 clinical patients)
master = decoded.merge(df_dig_slim, on='FICHA', how='left')

print(f"Master table: {master.shape[0]} patients, {master.shape[1]} columns")
print(f"With digital data: {master['Slide_ID'].notna().sum()} patients")
print(f"Without digital data: {master['Slide_ID'].isna().sum()} patients")

# ── Build Excel ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════
# SHEET 1: MASTER DATA TABLE
# ════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "All_Patients_Data"

# Title row
ws.merge_cells(f"A1:{get_column_letter(len(master.columns))}1")
title_cell = ws["A1"]
title_cell.value = (f"COMPLETE DATA TABLE — Lung Adenocarcinoma Study  |  "
                    f"n = {len(master)} patients  |  "
                    f"Digital pathology data: {master['Slide_ID'].notna().sum()} / {len(master)} patients")
title_cell.font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# Column headers
clinical_end = master.columns.tolist().index('Slide_ID')  # where digital starts

for c_idx, col_name in enumerate(master.columns, 1):
    cell = ws.cell(row=2, column=c_idx, value=col_name)
    # Clinical cols in blue, digital cols in green
    if c_idx <= clinical_end:
        cell.fill = PatternFill("solid", fgColor=MID_BLUE)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    else:
        cell.fill = PatternFill("solid", fgColor="375623")  # dark green
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
ws.row_dimensions[2].height = 36

# Data rows
for r_idx, (_, row) in enumerate(master.iterrows(), 3):
    has_digital = pd.notna(row.get('Slide_ID', None))
    row_bg = VERY_LIGHT if r_idx % 2 == 0 else WHITE

    for c_idx, col_name in enumerate(master.columns, 1):
        val = row[col_name]
        if pd.isna(val) or val == 'nan':
            val = ""
        elif isinstance(val, float) and not np.isnan(val) and val == int(val):
            val = int(val)

        cell = ws.cell(row=r_idx, column=c_idx, value=val)

        # Digital columns get yellow background if data present, orange if missing
        if c_idx > clinical_end:
            if has_digital and val != "":
                cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
            elif not has_digital:
                cell.fill = PatternFill("solid", fgColor=ORANGE)
            else:
                cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
        else:
            cell.fill = PatternFill("solid", fgColor=row_bg)

        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

# Freeze top 2 rows + column A
ws.freeze_panes = "B3"

# Auto column widths
for c_idx, col_name in enumerate(master.columns, 1):
    col_letter = get_column_letter(c_idx)
    max_len = max(len(str(col_name)), 8)
    for r_idx in range(3, min(len(master)+3, 20)):  # sample first 17 rows
        val = ws.cell(row=r_idx, column=c_idx).value
        if val:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

# ════════════════════════════════════════════════════════════════════════
# SHEET 2: VARIABLE LEGEND
# ════════════════════════════════════════════════════════════════════════
ws_leg = wb.create_sheet("Variable_Legend")
ws_leg.merge_cells("A1:C1")
t = ws_leg["A1"]; t.value = "VARIABLE LEGEND — Code Meanings"
t.font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_leg.row_dimensions[1].height = 28

for c, h in enumerate(["Variable", "Code / Value", "Meaning"], 1):
    cell = ws_leg.cell(row=2, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

legend_rows = [
    # Variable, Code, Meaning
    ("Patient_sex", "M", "Male"),
    ("Patient_sex", "F", "Female"),
    ("", "", ""),
    ("Status", "alive", "Alive at last contact (censored)"),
    ("Status", "dead", "Died (event)"),
    ("", "", ""),
    ("STAS", "No", "No spread through air spaces"),
    ("STAS", "Yes", "STAS positive"),
    ("STAS", "unknown", "STAS status indeterminate"),
    ("", "", ""),
    ("Lymphovascular_invasion", "No", "No lymphovascular invasion"),
    ("Lymphovascular_invasion", "Yes", "LVI positive"),
    ("", "", ""),
    ("Visceral_pleural_invasion", "no", "No visceral pleural invasion"),
    ("Visceral_pleural_invasion", "yes", "VPI positive"),
    ("", "", ""),
    ("pN", "N0", "No nodal metastasis"),
    ("pN", "N1", "Ipsilateral peribronchial / hilar LN metastasis"),
    ("pN", "N2", "Ipsilateral mediastinal LN metastasis"),
    ("", "", ""),
    ("Stage_group", "1A1", "pT1mi or T1a N0 M0"),
    ("Stage_group", "IA2", "T1b N0 M0"),
    ("Stage_group", "IA3", "T1c N0 M0"),
    ("Stage_group", "IB",  "T2a N0 M0"),
    ("Stage_group", "IIA", "T2b N0 M0"),
    ("Stage_group", "IIB", "T1-T2 N1 M0 or T3 N0 M0"),
    ("Stage_group", "IIIA","T3 N1, T4 N0/N1, or T1-T2 N2 M0"),
    ("Stage_group", "IIIB","T3-T4 N2 M0 or any T N3 M0"),
    ("", "", ""),
    ("Tumor_predominant_pattern", "acinar", "Acinar growth pattern predominant"),
    ("Tumor_predominant_pattern", "solid",  "Solid growth pattern predominant"),
    ("Tumor_predominant_pattern", "lepidic","Lepidic growth pattern predominant"),
    ("Tumor_predominant_pattern", "papilar","Papillary growth pattern predominant"),
    ("Tumor_predominant_pattern", "micropapilar","Micropapillary growth pattern predominant"),
    ("Tumor_predominant_pattern", "NA",    "Pattern not assessable"),
    ("", "", ""),
    ("PDL1_IHC", "<1",        "PD-L1 expression < 1% (negative)"),
    ("PDL1_IHC", "=>1",       "PD-L1 expression ≥ 1% (positive)"),
    ("PDL1_IHC", "not tested","PD-L1 IHC not performed"),
    ("", "", ""),
    ("Vimentin_IHC", "negative",  "Vimentin IHC negative"),
    ("Vimentin_IHC", "positive",  "Vimentin IHC positive"),
    ("Vimentin_IHC", "not tested","Vimentin IHC not performed"),
    ("", "", ""),
    ("Molecular_results", "not done",            "Molecular testing not performed"),
    ("Molecular_results", "no mutation",          "Tested — no driver mutation detected"),
    ("Molecular_results", "EGFR exon 19 deletion","EGFR exon 19 deletion"),
    ("Molecular_results", "EGFR L858R",           "EGFR L858R substitution"),
    ("Molecular_results", "EGFR other",           "Other EGFR mutation"),
    ("Molecular_results", "KRAS G12C mutation",   "KRAS G12C mutation"),
    ("Molecular_results", "ALK rearrangement",    "ALK rearrangement"),
    ("Molecular_results", "ALK fusion",           "ALK fusion"),
    ("Molecular_results", "ROS1 rearrangement",   "ROS1 rearrangement"),
    ("Molecular_results", "RET fusion",           "RET fusion"),
    ("", "", ""),
    ("Surgical_procedure", "lobectomy",   "Lobectomy"),
    ("Surgical_procedure", "pneumectomy", "Pneumonectomy"),
    ("Surgical_procedure", "bilobectomy", "Bilobectomy"),
    ("", "", ""),
    ("Margins", "free",     "Surgical margins tumour-free"),
    ("Margins", "positive", "Tumour at surgical margin"),
    ("", "", ""),
    ("=== DIGITAL PATHOLOGY COLUMNS ===", "", ""),
    ("Slide_ID",                        "e.g. R1-S7", "Scanner slide position identifier"),
    ("Lymp_density_ROI1 (cells/mm2)",   "numeric", "Lymphocyte density in Region of Interest 1"),
    ("Lymp_density_ROI2 (cells/mm2)",   "numeric", "Lymphocyte density in Region of Interest 2"),
    ("Lymp_density_MEAN (cells/mm2)",   "numeric", "Mean lymphocyte density (ROI1 + ROI2) / 2"),
    ("Tumor_density_ROI1 (cells/mm2)",  "numeric", "Tumour cell density in ROI 1"),
    ("Tumor_density_ROI2 (cells/mm2)",  "numeric", "Tumour cell density in ROI 2"),
    ("Tumor_density_MEAN (cells/mm2)",  "numeric", "Mean tumour cell density (ROI1 + ROI2) / 2"),
    ("Tumor_Lymp_ratio",                "numeric", "Tumour:lymphocyte ratio = Tumor_MEAN / Lymp_MEAN"),
    ("", "", ""),
    ("NOTE: Orange cells = no digital data", "(31 patients not mapped)", "Slide ID could not be linked to clinical record"),
    ("NOTE: Green cells = digital data present", "(141 patients)", "Used in Chapter 4 analysis"),
]

for r_idx, (var, code, meaning) in enumerate(legend_rows, 3):
    bg = VERY_LIGHT if r_idx % 2 == 0 else WHITE
    if var.startswith("==="):
        # Section header
        ws_leg.merge_cells(f"A{r_idx}:C{r_idx}")
        cell = ws_leg.cell(row=r_idx, column=1, value=var)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill = PatternFill("solid", fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    elif var.startswith("NOTE"):
        ws_leg.merge_cells(f"A{r_idx}:C{r_idx}") if code == "" else None
        for c_idx, val in enumerate([var, code, meaning], 1):
            cell = ws_leg.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(italic=True, name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor=YELLOW)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border()
    elif var == "" and code == "":
        ws_leg.row_dimensions[r_idx].height = 6
    else:
        for c_idx, val in enumerate([var, code, meaning], 1):
            cell = ws_leg.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=10, bold=(c_idx==1 and var!=""))
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border()

ws_leg.column_dimensions['A'].width = 35
ws_leg.column_dimensions['B'].width = 25
ws_leg.column_dimensions['C'].width = 55

# ════════════════════════════════════════════════════════════════════════
# SHEET 3: DIGITAL-ONLY SUBSET (141 mapped cases, numeric codes for SPSS import)
# ════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("SPSS_Import_141cases")
ws3.merge_cells("A1:Z1")
t3 = ws3["A1"]
t3.value = ("SPSS-READY: 141 mapped cases with digital pathology data  |  "
            "Numeric codes preserved for SPSS import  |  See legend sheet for code meanings")
t3.font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
t3.fill = PatternFill("solid", fgColor=DARK_BLUE)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

# Use original numeric codes for SPSS compatibility
df_spss2 = df_spss.copy()
df_spss2['Surgery_date'] = pd.to_datetime(df_spss2['Surgery_date'])
df_spss2['dateofdeathorlastcontact'] = pd.to_datetime(df_spss2['dateofdeathorlastcontact'])
df_spss2['Survival_months'] = ((df_spss2['dateofdeathorlastcontact'] - df_spss2['Surgery_date']).dt.days / 30.44).round(1)

spss_cols = ['FICHA', 'Patient_sex', 'Patient_age', 'Survival_months', 'Status',
             'STAS', 'Lymphovascular_Invasion', 'Visceral_Pleural_Invasion',
             'pT', 'pN', 'Stage_group', 'Tumor_predominant_pattern',
             'Number_lymph_nodes_submitted', 'Number_metastatic_lymph_nodes',
             'PDL1_IHQ', 'Vimentine_IHQ', 'Molecular_results', 'Tumor_size']

df_spss_slim = df_spss2[spss_cols].copy()

# Merge with digital data
df_spss_merged = df_spss_slim.merge(df_dig_slim, on='FICHA', how='inner')
print(f"SPSS import sheet: {len(df_spss_merged)} cases")

all_cols = list(df_spss_merged.columns)
for c_idx, col in enumerate(all_cols, 1):
    cell = ws3.cell(row=2, column=c_idx, value=col)
    is_digital = col in ['Slide_ID'] or 'density' in col.lower() or 'ratio' in col.lower() or 'Lymp' in col or 'Tumor_d' in col or 'Tumor_L' in col
    cell.fill = PatternFill("solid", fgColor="375623" if is_digital else MID_BLUE)
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
ws3.row_dimensions[2].height = 36

for r_idx, (_, row) in enumerate(df_spss_merged.iterrows(), 3):
    bg = VERY_LIGHT if r_idx % 2 == 0 else WHITE
    for c_idx, col in enumerate(all_cols, 1):
        val = row[col]
        if pd.isna(val): val = ""
        elif isinstance(val, float) and val == int(val): val = int(val)
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        is_digital = col in ['Slide_ID'] or 'density' in col.lower() or 'ratio' in col.lower() or 'Lymp' in col or 'Tumor_d' in col or 'Tumor_L' in col
        cell.fill = PatternFill("solid", fgColor=GREEN_FILL if is_digital else bg)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

ws3.freeze_panes = "B3"
for c_idx in range(1, len(all_cols)+1):
    ws3.column_dimensions[get_column_letter(c_idx)].width = 18

# Save
out_path = r'C:\Users\musha\Desktop\QuPath\Complete_Data_Table.xlsx'
wb.save(out_path)
print(f"\nSaved: {out_path}")
print("Sheets:")
for s in wb.sheetnames:
    print(f"  - {s}")
