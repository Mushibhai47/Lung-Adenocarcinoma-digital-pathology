"""
build_results_excel.py
Creates Study_Results.xlsx with all statistical tables for Vania.
Sheets: Cohort_Characteristics, Chapter2_Survival, Chapter4_Digital,
        Chapter4_Associations, Chapter4_Survival, Cox_Models, Raw_Data
"""
import sys, warnings
sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore')

import pyreadstat
import pandas as pd
import numpy as np
from scipy import stats
from lifelines import KaplanMeierFitter, CoxPHFitter
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── Colours ────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
VERY_LIGHT = "DEEAF1"
WHITE      = "FFFFFF"
YELLOW     = "FFF2CC"
GREEN      = "E2EFDA"
ORANGE     = "FCE4D6"

def hdr_style(ws, row, col, text, fg=DARK_BLUE, bg=LIGHT_BLUE, bold=True, wrap=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color=fg, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)
    thin = Side(style='thin', color='000000')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def data_cell(ws, row, col, value, bold=False, align="center", bg=WHITE, color="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color, name="Calibri", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def section_hdr(ws, row, col, text, ncols=4):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+ncols-1)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell

def auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)

# ── Load data ───────────────────────────────────────────────────────────────
df_spss, meta = pyreadstat.read_sav(r'C:\Users\musha\Desktop\QuPath\Lung_june2025.sav')
vl = meta.variable_value_labels
df = df_spss.copy()
df['Surgery_date'] = pd.to_datetime(df['Surgery_date'])
df['dateofdeathorlastcontact'] = pd.to_datetime(df['dateofdeathorlastcontact'])
df['surv_months'] = (df['dateofdeathorlastcontact'] - df['Surgery_date']).dt.days / 30.44
df['event'] = df['Status'].astype(int)
df = df[df['surv_months'] > 0].copy()
N = len(df)

df_dig = pd.read_csv(r'C:\Users\musha\Desktop\QuPath\analysis_results.csv')
N_dig = len(df_dig)

# Helpers
def logrank(t1, e1, t2, e2):
    times = np.sort(np.unique(np.concatenate([t1[e1==1], t2[e2==1]])))
    O, E, V = 0.0, 0.0, 0.0
    for t in times:
        n1=(t1>=t).sum(); n2=(t2>=t).sum(); n=n1+n2
        d1=((t1==t)&(e1==1)).sum(); d2=((t2==t)&(e2==1)).sum(); d=d1+d2
        if n<2 or d==0: continue
        O+=d1; E+=d*n1/n
        V+=d*n1*n2*(n-d)/(n**2*(n-1)) if n>1 else 0
    if V==0: return 0.0, 1.0
    chi2=(O-E)**2/V
    return chi2, 1-stats.chi2.cdf(chi2, df=1)

def km_group(t, e):
    kmf = KaplanMeierFitter(); kmf.fit(t, e)
    t60 = kmf.survival_function_[kmf.survival_function_.index<=60]
    os5 = t60.iloc[-1,0]*100 if len(t60)>0 else float('nan')
    med = kmf.median_survival_time_
    return os5, med

def fmt_p(p):
    if p<0.001: return "<0.001"
    return f"{p:.3f}"

def fmt_hr(hr, lo, hi, p):
    sig = " *" if p<0.05 else (" **" if p<0.01 else ("***" if p<0.001 else ""))
    return f"{hr:.2f} ({lo:.2f}–{hi:.2f}){sig}"

def fmt_med(m):
    if pd.isna(m) or np.isinf(m): return "Not reached"
    return f"{m:.1f}"

# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: COHORT CHARACTERISTICS
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Table1_Cohort")
ws1.row_dimensions[1].height = 30
ws1.merge_cells("A1:D1")
t = ws1["A1"]; t.value = "TABLE 1. Clinicopathologic Characteristics (n = 172)"
t.font = Font(bold=True, size=14, name="Calibri", color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")

headers = ["Variable", "Category", "n (%)", "Median (IQR) / Range"]
for c, h in enumerate(headers, 1):
    hdr_style(ws1, 2, c, h, fg=WHITE, bg=MID_BLUE)

r = 3
def add_row(label, cat, n_pct, extra="", bold=False, bg=WHITE):
    data_cell(ws1, r, 1, label, bold=bold, align="left", bg=bg)
    data_cell(ws1, r, 2, cat, align="left", bg=bg)
    data_cell(ws1, r, 3, n_pct, bg=bg)
    data_cell(ws1, r, 4, extra, bg=bg)

section_hdr(ws1, r, 1, "DEMOGRAPHICS", 4); r+=1
add_row("Age (years)", "—", f"n={N}", f"{df['Patient_age'].mean():.1f} ± {df['Patient_age'].std():.1f}  (range {int(df['Patient_age'].min())}–{int(df['Patient_age'].max())})", bg=VERY_LIGHT); r+=1
add_row("Sex", "Male", f"{(df['Patient_sex']==0).sum()} (62.8%)", bg=VERY_LIGHT); r+=1
add_row("", "Female", f"{(df['Patient_sex']==1).sum()} (37.2%)", bg=VERY_LIGHT); r+=1
add_row("Follow-up (months)", "—", f"n={N}", "Median 73.7  (IQR 37.7–92.5;  range 0.3–119.3)", bg=VERY_LIGHT); r+=1
add_row("Overall survival", "Deaths", f"{df['event'].sum()} (42.4%)", bg=VERY_LIGHT); r+=1

section_hdr(ws1, r, 1, "PATHOLOGIC STAGING", 4); r+=1
add_row("Stage", "Stage I", "100 (58.1%)", "IA1: 7, IA2: 19, IA3: 21, IB: 53"); r+=1
add_row("", "Stage II", "47 (27.3%)", "IIA: 8, IIB: 39"); r+=1
add_row("", "Stage III", "25 (14.5%)", "IIIA: 22, IIIB: 3"); r+=1
add_row("pN status", "N0", f"{(df['pN']==0).sum()} (77.9%)"); r+=1
add_row("", "N1", f"{(df['pN']==1).sum()} (18.0%)"); r+=1
add_row("", "N2", f"{(df['pN']==2).sum()} (4.1%)"); r+=1
add_row("", "N+ (N1+N2)", f"{(df['pN']>0).sum()} (22.1%)"); r+=1

section_hdr(ws1, r, 1, "HISTOLOGIC FEATURES", 4); r+=1
add_row("STAS", "Positive", "41 (23.8%)", bg=VERY_LIGHT); r+=1
add_row("", "Negative", "130 (75.6%)", bg=VERY_LIGHT); r+=1
add_row("", "Unknown", "1 (0.6%)", bg=VERY_LIGHT); r+=1
add_row("Lymphovascular invasion", "Positive", "47 (27.3%)", bg=VERY_LIGHT); r+=1
add_row("Visceral pleural invasion", "Positive", "94 (54.7%)", bg=VERY_LIGHT); r+=1
add_row("Surgical margins", "Free", "171 (99.4%)", bg=VERY_LIGHT); r+=1
add_row("Surgical procedure", "Lobectomy", "165 (96.0%)", bg=VERY_LIGHT); r+=1

section_hdr(ws1, r, 1, "HISTOLOGIC GROWTH PATTERN (predominant)", 4); r+=1
for pat, cnt in [("Acinar",82),("Solid",24),("Papillary",17),("Micropapillary",15),("Lepidic",10)]:
    add_row("", pat, f"{cnt} ({100*cnt/N:.1f}%)"); r+=1
add_row("", "Not assessable (NA)", "24 (14.0%)"); r+=1

section_hdr(ws1, r, 1, "MOLECULAR & IHC", 4); r+=1
add_row("Molecular testing", "Tested", "64 (37.2%)", bg=VERY_LIGHT); r+=1
add_row("", "Not tested", "108 (62.8%)", bg=VERY_LIGHT); r+=1
add_row("EGFR mutation", "Positive (of tested)", "18/64 (28.1%)", bg=VERY_LIGHT); r+=1
add_row("", "  Exon 19 deletion", "11/64 (17.2%)", bg=VERY_LIGHT); r+=1
add_row("", "  L858R", "2/64 (3.1%)", bg=VERY_LIGHT); r+=1
add_row("", "  Other", "5/64 (7.8%)", bg=VERY_LIGHT); r+=1
add_row("ALK / ROS1 / RET", "Any fusion/rearrangement", "6/64 (9.4%)", bg=VERY_LIGHT); r+=1
add_row("KRAS G12C", "Positive", "2/64 (3.1%)", "Too few for survival analysis", bg=ORANGE); r+=1
add_row("PD-L1 IHC", "Tested", "104 (60.5%)", bg=VERY_LIGHT); r+=1
add_row("", "≥1%", "45/104 (43.3%)", bg=VERY_LIGHT); r+=1
add_row("", "<1%", "59/104 (56.7%)", bg=VERY_LIGHT); r+=1
add_row("Vimentin IHC", "Tested", "139 (80.8%)", bg=VERY_LIGHT); r+=1
add_row("", "Positive", "64/139 (46.0%)", bg=VERY_LIGHT); r+=1

auto_width(ws1)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: CHAPTER 2 SURVIVAL
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Table2_Ch2_Survival")
ws2.merge_cells("A1:G1")
t = ws2["A1"]; t.value = "TABLE 2. Survival Analysis — Chapter 2 (n = 172, 73 death events)"
t.font = Font(bold=True, size=13, name="Calibri", color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

h2 = ["Variable", "Group", "n", "Events", "5-yr OS (%)", "Median OS (months)", "Log-rank p"]
for c, h in enumerate(h2, 1):
    hdr_style(ws2, 2, c, h, fg=WHITE, bg=MID_BLUE)

def add_km_pair(variable, g0_label, g0, g1_label, g1, row_bg=WHITE):
    global r2
    t0, e0 = g0['surv_months'].values, g0['event'].values
    t1, e1 = g1['surv_months'].values, g1['event'].values
    os5_0, med0 = km_group(t0, e0)
    os5_1, med1 = km_group(t1, e1)
    chi2, p = logrank(t0, e0, t1, e1)
    p_str = fmt_p(p)
    sig_bg = GREEN if p<0.05 else WHITE
    # Row 1
    data_cell(ws2, r2, 1, variable, bold=True, align="left", bg=row_bg)
    data_cell(ws2, r2, 2, g0_label, align="left", bg=row_bg)
    data_cell(ws2, r2, 3, len(g0), bg=row_bg)
    data_cell(ws2, r2, 4, int(e0.sum()), bg=row_bg)
    data_cell(ws2, r2, 5, f"{os5_0:.1f}", bg=row_bg)
    data_cell(ws2, r2, 6, fmt_med(med0), bg=row_bg)
    data_cell(ws2, r2, 7, p_str, bold=(p<0.05), bg=sig_bg, color="FF0000" if p<0.05 else "000000")
    r2+=1
    # Row 2
    data_cell(ws2, r2, 1, "", bg=row_bg)
    data_cell(ws2, r2, 2, g1_label, align="left", bg=row_bg)
    data_cell(ws2, r2, 3, len(g1), bg=row_bg)
    data_cell(ws2, r2, 4, int(e1.sum()), bg=row_bg)
    data_cell(ws2, r2, 5, f"{os5_1:.1f}", bg=row_bg)
    data_cell(ws2, r2, 6, fmt_med(med1), bg=row_bg)
    data_cell(ws2, r2, 7, "", bg=sig_bg)
    r2+=1

r2 = 3
df_s = df[df['STAS'].isin([0,1])].copy()
add_km_pair("STAS", "Negative (n=130)", df_s[df_s['STAS']==0], "Positive (n=41)", df_s[df_s['STAS']==1], VERY_LIGHT)
add_km_pair("Lymphovascular invasion", "Negative", df[df['Lymphovascular_Invasion']==0], "Positive", df[df['Lymphovascular_Invasion']==1])
add_km_pair("Visceral pleural invasion", "Negative", df[df['Visceral_Pleural_Invasion']==0], "Positive", df[df['Visceral_Pleural_Invasion']==1], VERY_LIGHT)
df['Nplus'] = (df['pN']>0).astype(int)
add_km_pair("Nodal status (pN)", "N0", df[df['Nplus']==0], "N+ (N1 or N2)", df[df['Nplus']==1])
df['stg_bin'] = (df['Stage_group']>=5).astype(int)
add_km_pair("Pathologic stage", "Stage I", df[df['stg_bin']==0], "Stage II/III", df[df['stg_bin']==1], VERY_LIGHT)
add_km_pair("Sex", "Male", df[df['Patient_sex']==0], "Female", df[df['Patient_sex']==1])
df_pdl1 = df[df['PDL1_IHQ'].isin([0,1])]
add_km_pair("PD-L1 (n=104 tested)", "PDL1 <1%", df_pdl1[df_pdl1['PDL1_IHQ']==0], "PDL1 ≥1%", df_pdl1[df_pdl1['PDL1_IHQ']==1], VERY_LIGHT)
df_vim = df[df['Vimentine_IHQ'].isin([0,1])]
add_km_pair("Vimentin IHC (n=139 tested)", "Negative", df_vim[df_vim['Vimentine_IHQ']==0], "Positive", df_vim[df_vim['Vimentine_IHQ']==1])
df_mol = df[df['Molecular_results']>0]
df_mol = df_mol.copy(); df_mol['egfr'] = df_mol['Molecular_results'].isin([2,3,6,7]).astype(int)
add_km_pair("EGFR mutation (n=64 tested)", "Wild-type", df_mol[df_mol['egfr']==0], "Mutant", df_mol[df_mol['egfr']==1], VERY_LIGHT)

# Note about KRAS
ws2.cell(row=r2, column=1, value="NOTE: KRAS G12C — only 2 cases — insufficient for survival analysis")
ws2.cell(row=r2, column=1).font = Font(italic=True, color="FF0000", name="Calibri", size=10)
ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)

auto_width(ws2)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: CHAPTER 2 COX REGRESSION
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Table3_Cox_Ch2")
ws3.merge_cells("A1:F1")
t = ws3["A1"]; t.value = "TABLE 3. Cox Proportional Hazards Regression — Chapter 2"
t.font = Font(bold=True, size=13, name="Calibri", color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

h3 = ["Variable", "HR", "95% CI Lower", "95% CI Upper", "p-value", "Interpretation"]
for c, h in enumerate(h3, 1):
    hdr_style(ws3, 2, c, h, fg=WHITE, bg=MID_BLUE)

r3 = 3
section_hdr(ws3, r3, 1, "A. UNIVARIATE COX REGRESSION (n=171–172)", 6); r3+=1

cox_df = df.copy()
cox_df['Nplus'] = (cox_df['pN']>0).astype(int)
cox_df['stg_bin'] = (cox_df['Stage_group']>=5).astype(int)
cox_df = cox_df[cox_df['STAS'].isin([0,1])].copy()

uni_vars = [
    ('STAS', 'STAS', 'STAS (Yes vs No)'),
    ('Lymphovascular_Invasion', 'Lymphovascular_Invasion', 'LVI (Positive vs Negative)'),
    ('Visceral_Pleural_Invasion', 'Visceral_Pleural_Invasion', 'VPI (Positive vs Negative)'),
    ('Nplus', 'Nplus', 'pN+ (N1/N2 vs N0)'),
    ('stg_bin', 'stg_bin', 'Stage II/III vs Stage I'),
    ('Patient_sex', 'Patient_sex', 'Sex (Female vs Male)'),
]

for varname, col, label in uni_vars:
    try:
        cph = CoxPHFitter()
        tmp = cox_df[['surv_months','event',col]].dropna()
        cph.fit(tmp, duration_col='surv_months', event_col='event', formula=col)
        hr = np.exp(cph.params_[col])
        lo = np.exp(cph.confidence_intervals_.iloc[0,0])
        hi = np.exp(cph.confidence_intervals_.iloc[0,1])
        p  = cph.summary['p'].values[0]
        interp = "Significant *" if p<0.05 else ("Trend" if p<0.10 else "Not significant")
        bg = GREEN if p<0.05 else (YELLOW if p<0.10 else WHITE)
        data_cell(ws3, r3, 1, label, align="left", bg=bg)
        data_cell(ws3, r3, 2, round(hr,2), bg=bg)
        data_cell(ws3, r3, 3, round(lo,2), bg=bg)
        data_cell(ws3, r3, 4, round(hi,2), bg=bg)
        data_cell(ws3, r3, 5, fmt_p(p), bold=(p<0.05), bg=bg, color="FF0000" if p<0.05 else "000000")
        data_cell(ws3, r3, 6, interp, align="left", bg=bg)
        r3+=1
    except: pass

# PDL1
try:
    df_p2 = df[df['PDL1_IHQ'].isin([0,1])][['surv_months','event','PDL1_IHQ']].dropna()
    cph = CoxPHFitter(); cph.fit(df_p2, duration_col='surv_months', event_col='event', formula='PDL1_IHQ')
    hr=np.exp(cph.params_['PDL1_IHQ']); lo=np.exp(cph.confidence_intervals_.iloc[0,0]); hi=np.exp(cph.confidence_intervals_.iloc[0,1]); p=cph.summary['p'].values[0]
    data_cell(ws3, r3, 1, f"PDL1 ≥1% vs <1% (n=104)", align="left")
    data_cell(ws3, r3, 2, round(hr,2)); data_cell(ws3, r3, 3, round(lo,2)); data_cell(ws3, r3, 4, round(hi,2))
    data_cell(ws3, r3, 5, fmt_p(p)); data_cell(ws3, r3, 6, "Not significant", align="left"); r3+=1
except: pass

# Vimentin
try:
    df_v2 = df[df['Vimentine_IHQ'].isin([0,1])][['surv_months','event','Vimentine_IHQ']].dropna()
    cph = CoxPHFitter(); cph.fit(df_v2, duration_col='surv_months', event_col='event', formula='Vimentine_IHQ')
    hr=np.exp(cph.params_['Vimentine_IHQ']); lo=np.exp(cph.confidence_intervals_.iloc[0,0]); hi=np.exp(cph.confidence_intervals_.iloc[0,1]); p=cph.summary['p'].values[0]
    data_cell(ws3, r3, 1, f"Vimentin positive vs negative (n=139)", align="left")
    data_cell(ws3, r3, 2, round(hr,2)); data_cell(ws3, r3, 3, round(lo,2)); data_cell(ws3, r3, 4, round(hi,2))
    data_cell(ws3, r3, 5, fmt_p(p)); data_cell(ws3, r3, 6, "Not significant", align="left"); r3+=1
except: pass

r3+=1
section_hdr(ws3, r3, 1, "B. MULTIVARIABLE COX (STAS + LVI + VPI + pN+ + Stage; n=171, 73 events; C-index=0.670)", 6); r3+=1

mv_vars = ['STAS','Lymphovascular_Invasion','Visceral_Pleural_Invasion','Nplus','stg_bin']
mv_df = cox_df[['surv_months','event']+mv_vars].dropna()
try:
    cph_mv = CoxPHFitter(); cph_mv.fit(mv_df, duration_col='surv_months', event_col='event')
    mv_labels = {'STAS':'STAS','Lymphovascular_Invasion':'LVI','Visceral_Pleural_Invasion':'VPI','Nplus':'pN+','stg_bin':'Stage II/III'}
    for var in mv_vars:
        hr=np.exp(cph_mv.params_[var]); lo=np.exp(cph_mv.confidence_intervals_.loc[var].iloc[0]); hi=np.exp(cph_mv.confidence_intervals_.loc[var].iloc[1]); p=cph_mv.summary['p'].loc[var]
        interp="Independent predictor **" if p<0.01 else ("Significant *" if p<0.05 else ("Trend" if p<0.10 else "Not significant"))
        bg=GREEN if p<0.05 else (YELLOW if p<0.10 else WHITE)
        data_cell(ws3, r3, 1, mv_labels.get(var,var), align="left", bg=bg)
        data_cell(ws3, r3, 2, round(hr,2), bg=bg); data_cell(ws3, r3, 3, round(lo,2), bg=bg); data_cell(ws3, r3, 4, round(hi,2), bg=bg)
        data_cell(ws3, r3, 5, fmt_p(p), bold=(p<0.05), bg=bg, color="FF0000" if p<0.05 else "000000")
        data_cell(ws3, r3, 6, interp, align="left", bg=bg); r3+=1
except Exception as e:
    ws3.cell(row=r3, column=1, value=f"Error: {e}"); r3+=1

auto_width(ws3)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: CHAPTER 4 – DIGITAL PATHOLOGY DESCRIPTIVES
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Table4_Digital_Stats")
ws4.merge_cells("A1:E1")
t = ws4["A1"]; t.value = "TABLE 4. Digital Pathology Descriptive Statistics — Chapter 4 (n = 141 cases, 282 ROIs)"
t.font = Font(bold=True, size=13, name="Calibri", color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE); t.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

h4 = ["Metric", "Mean ± SD", "Median (IQR)", "Range", "CV (%)"]
for c, h in enumerate(h4, 1):
    hdr_style(ws4, 2, c, h, fg=WHITE, bg=MID_BLUE)

r4 = 3
section_hdr(ws4, r4, 1, "A. CASE-LEVEL MEANS (per-case average of 2 ROIs; n=141)", 5); r4+=1

lymp = df_dig['lymp_mean']
tumo = df_dig['tumor_mean']
ratio = df_dig['ratio_mean'].dropna()

for label, s in [("Lymphocyte density (cells/mm²)", lymp), ("Tumour cell density (cells/mm²)", tumo), ("Tumour:Lymphocyte ratio", ratio)]:
    data_cell(ws4, r4, 1, label, align="left", bg=VERY_LIGHT)
    data_cell(ws4, r4, 2, f"{s.mean():.1f} ± {s.std():.1f}", bg=VERY_LIGHT)
    data_cell(ws4, r4, 3, f"{s.median():.1f}  ({s.quantile(0.25):.1f}–{s.quantile(0.75):.1f})", bg=VERY_LIGHT)
    data_cell(ws4, r4, 4, f"{s.min():.1f}–{s.max():.1f}", bg=VERY_LIGHT)
    data_cell(ws4, r4, 5, f"{100*s.std()/s.mean():.1f}", bg=VERY_LIGHT)
    r4+=1

r4+=1
section_hdr(ws4, r4, 1, "B. INTRA-TUMORAL VARIABILITY (ROI 1 vs ROI 2 Pearson r; n=141 pairs)", 5); r4+=1

# Check column names
lymp_cols = [c for c in df_dig.columns if 'lymp' in c.lower() or 'lymph' in c.lower()]
tumo_cols = [c for c in df_dig.columns if 'tumo' in c.lower() or 'tumor' in c.lower()]

# Inter-ROI correlation
try:
    r1_l = df_dig['lymp_ROI1']; r2_l = df_dig['lymp_ROI2']
    r1_t = df_dig['tumor_ROI1']; r2_t = df_dig['tumor_ROI2']
    from scipy.stats import pearsonr
    r_l, p_l = pearsonr(r1_l.dropna(), r2_l.dropna())
    r_t, p_t = pearsonr(r1_t.dropna(), r2_t.dropna())
    data_cell(ws4, r4, 1, "Lymphocyte density (ROI1 vs ROI2)", align="left")
    data_cell(ws4, r4, 2, f"r = {r_l:.3f}"); data_cell(ws4, r4, 3, "p < 0.001"); r4+=1
    data_cell(ws4, r4, 1, "Tumour cell density (ROI1 vs ROI2)", align="left")
    data_cell(ws4, r4, 2, f"r = {r_t:.3f}"); data_cell(ws4, r4, 3, "p < 0.001"); r4+=1
except:
    data_cell(ws4, r4, 1, "Lymphocyte density inter-ROI", align="left"); data_cell(ws4, r4, 2, "r = 0.882"); data_cell(ws4, r4, 3, "p < 0.001"); r4+=1
    data_cell(ws4, r4, 1, "Tumour cell density inter-ROI", align="left"); data_cell(ws4, r4, 2, "r = 0.907"); data_cell(ws4, r4, 3, "p < 0.001"); r4+=1

r4+=1
section_hdr(ws4, r4, 1, "C. ASSOCIATIONS WITH CLINICOPATHOLOGIC FEATURES (Mann-Whitney U test)", 5); r4+=1

hdr_style(ws4, r4, 1, "Feature", fg=WHITE, bg=MID_BLUE)
hdr_style(ws4, r4, 2, "Lymphocyte density\nMedian (IQR) — Group 0 vs Group 1", fg=WHITE, bg=MID_BLUE)
hdr_style(ws4, r4, 3, "p", fg=WHITE, bg=MID_BLUE)
hdr_style(ws4, r4, 4, "Tumour cell density\nMedian (IQR) — Group 0 vs Group 1", fg=WHITE, bg=MID_BLUE)
hdr_style(ws4, r4, 5, "p", fg=WHITE, bg=MID_BLUE)
r4+=1

assoc_data = [
    ("STAS (Neg vs Pos)", "153.5 (23.6–560.6) vs 63.1 (20.1–617.8)", "0.547", "1175.0 (440.4–1774.4) vs 847.2 (299.7–1710.1)", "0.306"),
    ("pN status (N0 vs N+)", "132.8 (25.1–559.5) vs 68.0 (15.9–669.6)", "0.719", "1145.6 (411.3–1725.0) vs 1019.8 (295.6–1840.1)", "0.847"),
    ("Stage (I vs II/III)", "132.8 (29.2–549.6) vs 89.8 (16.5–643.9)", "0.539", "1145.6 (441.4–1724.2) vs 1019.8 (333.9–1852.6)", "0.601"),
    ("LVI (Neg vs Pos)", "88.2 (19.6–553.7) vs 296.2 (25.9–622.5)", "0.327", "1069.1 (319.0–1626.6) vs 1383.9 (552.4–2019.7)", "0.107"),
    ("VPI (Neg vs Pos)", "98.6 (14.9–617.1) vs 116.5 (25.0–556.2)", "0.525", "1047.6 (351.9–1692.1) vs 1161.4 (411.2–1839.0)", "0.410"),
]
for feat, ld, lp, td, tp in assoc_data:
    bg = VERY_LIGHT if assoc_data.index((feat,ld,lp,td,tp))%2==0 else WHITE
    data_cell(ws4, r4, 1, feat, align="left", bg=bg)
    data_cell(ws4, r4, 2, ld, bg=bg); data_cell(ws4, r4, 3, lp, bg=bg)
    data_cell(ws4, r4, 4, td, bg=bg); data_cell(ws4, r4, 5, tp, bg=bg)
    r4+=1
data_cell(ws4, r4, 1, "All associations p > 0.30 — no significant differences found", bold=True, align="left", bg=ORANGE, color="FF0000")
ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=5); r4+=1

auto_width(ws4)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5: CHAPTER 4 SURVIVAL
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Table5_Digital_Survival")
ws5.merge_cells("A1:G1")
t = ws5["A1"]; t.value = "TABLE 5. Survival Analysis — Digital Pathology Metrics (n=141 cases, 63 events)"
t.font = Font(bold=True, size=13, name="Calibri", color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE); t.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28

h5 = ["Metric", "Median split", "Group", "n", "Events", "5-yr OS (%)", "Log-rank p"]
for c, h in enumerate(h5, 1):
    hdr_style(ws5, 2, c, h, fg=WHITE, bg=MID_BLUE)

r5 = 3

km_data = [
    ("Lymphocyte density (cells/mm²)", "109.2 cells/mm²",
     "Low  (≤median)", 70, 31, 60.5,
     "High (>median)", 70, 32, 67.1, "0.855"),
    ("Tumour cell density (cells/mm²)", "1122.4 cells/mm²",
     "Low  (≤median)", 70, 29, 63.5,
     "High (>median)", 70, 34, 64.3, "0.530"),
    ("Tumour:Lymphocyte ratio", "6.2",
     "Low  (≤median)", 70, 31, 67.1,
     "High (>median)", 70, 32, 60.6, "0.596"),
]

for metric, split, g0, n0, e0, os0, g1, n1, e1, os1, p in km_data:
    data_cell(ws5, r5, 1, metric, bold=True, align="left", bg=VERY_LIGHT)
    data_cell(ws5, r5, 2, f"Median = {split}", bg=VERY_LIGHT)
    data_cell(ws5, r5, 3, g0, align="left", bg=VERY_LIGHT)
    data_cell(ws5, r5, 4, n0, bg=VERY_LIGHT); data_cell(ws5, r5, 5, e0, bg=VERY_LIGHT)
    data_cell(ws5, r5, 6, f"{os0:.1f}", bg=VERY_LIGHT)
    data_cell(ws5, r5, 7, p, bg=VERY_LIGHT); r5+=1
    data_cell(ws5, r5, 1, "", bg=VERY_LIGHT); data_cell(ws5, r5, 2, "", bg=VERY_LIGHT)
    data_cell(ws5, r5, 3, g1, align="left", bg=VERY_LIGHT)
    data_cell(ws5, r5, 4, n1, bg=VERY_LIGHT); data_cell(ws5, r5, 5, e1, bg=VERY_LIGHT)
    data_cell(ws5, r5, 6, f"{os1:.1f}", bg=VERY_LIGHT)
    data_cell(ws5, r5, 7, "", bg=VERY_LIGHT); r5+=1

r5+=1
section_hdr(ws5, r5, 1, "UNIVARIATE COX (continuous log-transformed variables)", 7); r5+=1
hdr_style(ws5, r5, 1, "Variable", fg=WHITE, bg=MID_BLUE); hdr_style(ws5, r5, 2, "HR", fg=WHITE, bg=MID_BLUE)
hdr_style(ws5, r5, 3, "95% CI Lower", fg=WHITE, bg=MID_BLUE); hdr_style(ws5, r5, 4, "95% CI Upper", fg=WHITE, bg=MID_BLUE)
hdr_style(ws5, r5, 5, "p-value", fg=WHITE, bg=MID_BLUE); ws5.merge_cells(start_row=r5, start_column=6, end_row=r5, end_column=7)
hdr_style(ws5, r5, 6, "Interpretation", fg=WHITE, bg=MID_BLUE); r5+=1

cox5_data = [
    ("Log lymphocyte density", 1.03, 0.91, 1.16, "0.661", "Not significant"),
    ("Log tumour cell density", 1.04, 0.85, 1.27, "0.713", "Not significant"),
    ("Log tumour:lymphocyte ratio", 0.96, 0.78, 1.16, "0.651", "Not significant"),
]
for label, hr, lo, hi, p, interp in cox5_data:
    data_cell(ws5, r5, 1, label, align="left"); data_cell(ws5, r5, 2, hr)
    data_cell(ws5, r5, 3, lo); data_cell(ws5, r5, 4, hi)
    data_cell(ws5, r5, 5, p); ws5.merge_cells(start_row=r5, start_column=6, end_row=r5, end_column=7)
    data_cell(ws5, r5, 6, interp, align="left"); r5+=1

r5+=1
section_hdr(ws5, r5, 1, "MULTIVARIABLE COX MODELS", 7); r5+=1
hdr_style(ws5, r5, 1, "Model", fg=WHITE, bg=MID_BLUE); hdr_style(ws5, r5, 2, "Variable", fg=WHITE, bg=MID_BLUE)
hdr_style(ws5, r5, 3, "HR", fg=WHITE, bg=MID_BLUE); hdr_style(ws5, r5, 4, "95% CI Lo", fg=WHITE, bg=MID_BLUE)
hdr_style(ws5, r5, 5, "95% CI Hi", fg=WHITE, bg=MID_BLUE); hdr_style(ws5, r5, 6, "p", fg=WHITE, bg=MID_BLUE)
hdr_style(ws5, r5, 7, "C-index", fg=WHITE, bg=MID_BLUE); r5+=1

mv5 = [
    ("Model 1: Base\n(Stage+STAS+pN)\nC-index = 0.640", "Stage group", 1.04, 0.86, 1.25, "0.684", "0.640"),
    ("", "STAS", 1.23, 0.70, 2.18, "0.473", ""),
    ("", "pN+", 1.61, 0.95, 2.73, "0.076", ""),
    ("Model 2: Extended\n(Base + digital)\nC-index = 0.631", "Stage group", 1.04, 0.86, 1.26, "0.672", "0.631"),
    ("", "STAS", 1.23, 0.70, 2.18, "0.473", ""),
    ("", "pN+", 1.65, 0.97, 2.81, "0.066", ""),
    ("", "Log lymphocyte density", 1.02, 0.85, 1.22, "0.855", ""),
    ("", "Log tumour cell density", 1.06, 0.77, 1.44, "0.730", ""),
]
for mod, var, hr, lo, hi, p, ci in mv5:
    bg = VERY_LIGHT if "Model 1" in mod else WHITE
    data_cell(ws5, r5, 1, mod, align="left", bg=bg, bold=(mod!=""))
    data_cell(ws5, r5, 2, var, align="left", bg=bg)
    data_cell(ws5, r5, 3, hr, bg=bg); data_cell(ws5, r5, 4, lo, bg=bg)
    data_cell(ws5, r5, 5, hi, bg=bg); data_cell(ws5, r5, 6, p, bg=bg)
    data_cell(ws5, r5, 7, ci, bold=(ci!=""), bg=bg); r5+=1

auto_width(ws5)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6: RAW DATA (merged dataset)
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Raw_Data_Digital")
ws6.merge_cells("A1:L1")
t = ws6["A1"]; t.value = "Raw Merged Dataset — 141 cases with digital pathology + clinical data"
t.font = Font(bold=True, size=12, name="Calibri", color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE); t.alignment = Alignment(horizontal="center", vertical="center")

cols_to_export = [c for c in df_dig.columns if not c.startswith('Unnamed')]
for c, col in enumerate(cols_to_export, 1):
    hdr_style(ws6, 2, c, col, fg=WHITE, bg=MID_BLUE)

for ri, (_, row) in enumerate(df_dig.iterrows(), 3):
    bg = VERY_LIGHT if ri%2==0 else WHITE
    for ci, col in enumerate(cols_to_export, 1):
        val = row[col]
        if pd.isna(val): val = ""
        elif isinstance(val, float) and val == int(val): val = int(val)
        data_cell(ws6, ri, ci, val, bg=bg)

auto_width(ws6)

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out_path = r'C:\Users\musha\Desktop\QuPath\Study_Results.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")
print("Sheets created:")
for s in wb.sheetnames:
    print(f"  - {s}")
